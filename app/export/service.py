import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from app.db import supabase_client


def _apply_table_styles(ws) -> None:
	header_font = Font(bold=True)
	align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
	thin_border = Border(
		left=Side(style="thin"),
		right=Side(style="thin"),
		top=Side(style="thin"),
		bottom=Side(style="thin"),
	)
	for cell in ws[1]:
		cell.font = header_font
		cell.alignment = align_left
		cell.border = thin_border
	for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
		for cell in row:
			cell.alignment = align_left
			cell.border = thin_border


async def export_workouts_xlsx(user_id: str) -> bytes:
	workouts = await supabase_client.get(
		"workouts",
		{"user_id": f"eq.{user_id}", "order": "date.desc", "limit": "10000"},
	)

	wb = Workbook()
	ws = wb.active
	ws.title = "Тренировки"
	ws.append(["Дата", "Тип", "Описание", "Оценка", "Комментарий"])
	if workouts:
		for w in workouts:
			rating = w.get("rating")
			rating_display = f"⭐ {rating}/5" if rating else "—"
			ws.append(
				[
					w.get("date", ""),
					w.get("workout_type", ""),
					w.get("details", ""),
					rating_display,
					w.get("comment") or "—",
				]
			)
	_apply_table_styles(ws)
	ws.column_dimensions["A"].width = 14
	ws.column_dimensions["B"].width = 14
	ws.column_dimensions["C"].width = 70
	ws.column_dimensions["D"].width = 12
	ws.column_dimensions["E"].width = 30

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


async def export_nutrition_xlsx(user_id: str) -> bytes:
	meals = await supabase_client.get(
		"meals",
		{"user_id": f"eq.{user_id}", "order": "date.desc", "limit": "10000"},
	)
	plans = await supabase_client.get(
		"nutrition_plans",
		{"user_id": f"eq.{user_id}", "order": "created_at.desc", "limit": "200"},
	)

	wb = Workbook()

	ws_meals = wb.active
	ws_meals.title = "Приемы пищи"
	ws_meals.append(["Дата", "Описание", "Калории", "Белки", "Жиры", "Углеводы"])
	if meals:
		for m in meals:
			ws_meals.append(
				[
					m.get("date", ""),
					m.get("description", ""),
					m.get("calories", ""),
					m.get("proteins", ""),
					m.get("fats", ""),
					m.get("carbs", ""),
				]
			)
	_apply_table_styles(ws_meals)

	ws_plans = wb.create_sheet("Планы питания")
	ws_plans.append(["Дата создания", "Цель", "Ограничения", "Предпочтения", "Время готовки", "Бюджет", "Статус"])
	if plans:
		for p in plans:
			created_at = p.get("created_at")
			created_str = ""
			if isinstance(created_at, str):
				try:
					dt = datetime.fromisoformat(created_at.replace("Z", "+00:00"))
					created_str = dt.strftime("%d.%m.%Y %H:%M")
				except Exception:
					created_str = created_at
			status = "Активный" if p.get("is_active") else "Неактивный"
			ws_plans.append(
				[
					created_str,
					p.get("nutrition_goal", ""),
					p.get("dietary_restrictions", ""),
					p.get("meal_preferences", ""),
					p.get("cooking_time", ""),
					p.get("budget", ""),
					status,
				]
			)
	_apply_table_styles(ws_plans)

	buf = io.BytesIO()
	wb.save(buf)
	return buf.getvalue()


